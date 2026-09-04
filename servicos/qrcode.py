import qrcode
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,Alignment
from openpyxl import load_workbook
import os
import json

from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

from servicos.apontamento import localizar_odp

FONTE_PADRAO = Font(name="Arial", size=11)

ALINHAMENTO_PADRAO = Alignment(horizontal="center",vertical="center",wrap_text=True)

# O texto servirá como parâmetro no lugar de texto_qr onde ODP, data, turno e operador são parâmetros obrigatórios pois são eles que nos ajudarão a encontrar o caminho até ao arquivo, aba e linha onde está nossa odp
def gerar_qrcode(texto,odp,operador):

# Pegue o conteúdo recebido em odp e transforme-o em uma imagem de QR Code.
        qr = qrcode.make(texto)

        odp_limpo = odp.replace("/","-")

        operario = operador

 # Aqui você aproveitou a própria ODP para criar um nome único para a imagem       
        test = f"QR_{odp_limpo}_{operario}.png"

        caminho = os.path.join("temporario", test)

# Agora a imagem que está na variável qr é realmente gravada no computador
        qr.save(caminho)

# A função não retorna o QR Code em si. Ela retorna o endereço onde a imagem foi salva
        return caminho

def interpretar_qrcode(identificador,rastreabilidade):
        dados = rastreabilidade[identificador]

        print(f"ODP: {dados['odp']}")
        print(f"OPERADOR: {dados['operador']}")
        print(f"DATA: {dados['data']}")
        print(f"TURNO: {dados["turno"]}")

        return dados

def preencher_etiqueta(ordem,rastreabilidade):

        identificador = ordem["identificador"]

        dados_identificador = {
                "odp": ordem["odp"],
                "numero_pedido": ordem["numero_pedido"],
                "cliente": ordem["cliente"],
                "operador": ordem["operador"],
                "maquina": ordem["maquina"],
                "turno": ordem["turno"],
                "data": ordem["data"].strftime("%d%m%Y"),
        }

        rastreabilidade[identificador] = dados_identificador

# Carrega a planilha que queremos
        planilha = load_workbook("modelos/Modelo_Etiqueta_Luari.xlsx")

# Abre na 1º aba do arquivo
        aba = planilha.active

        aba["B6"] = ordem["odp"]
        aba["B6"].font = FONTE_PADRAO
        aba["B6"].alignment = ALINHAMENTO_PADRAO

        aba["B7"] = ordem["operador"]
        aba["B7"].font = FONTE_PADRAO
        aba["B7"].alignment = ALINHAMENTO_PADRAO

        aba["B8"] = ordem["maquina"]
        aba["B8"].font = FONTE_PADRAO
        aba["B8"].alignment = ALINHAMENTO_PADRAO

        aba["B9"] = ordem["data"]
        aba["B9"].number_format = "dd/mm/yy"
        aba["B9"].font = FONTE_PADRAO
        aba["B9"].alignment = ALINHAMENTO_PADRAO

# Aqui ele pega e transforma o texto em uma linguagem que o programa do QR code consiga ler,no caso formato de texto, esse comando ensure_ascii= False serve para caso seja escrito uma palavra com ~ ou ç o programa faça o texto ficar normal e não um conjunto de letras estranhas
        texto_qr = identificador

# Usa a função Gerar_qrcode para gerar o caminho da imagem PNG do QRcode. A odp aí é apenas pq a função gerar-qrcode exige esses 2 parâmetros
        caminho_qr = gerar_qrcode(texto_qr,ordem["odp"],ordem["operador"])

# Cria um objeto que pode ser inserido na planilha
        qr = Image(caminho_qr)

# Define os tamahos 406x240(valores do espaço onde o QR code está inserido)
        qr.height = 245
        qr.width = 245

# col define a coluna B,row a linha 11 e as terminações em Off são pra encostar certinho das bordas da coluna e célula especificado, sem espaçamento desnecessário
#         marcador = AnchorMarker(col = 1,colOff = 0, row = 10, rowOff = 0)

# # 
#         qr.anchor = OneCellAnchor(_from = marcador, ext = XDRPositiveSize2D(pixels_to_EMU(qr.width), pixels_to_EMU(qr.height)))

# Adiciona a imagem começando na célula especificada
        aba.add_image(qr,"B11")

        odp_limpo = ordem["odp"].replace("/","-")

        operario = ordem["operador"]

# Camiho da saída do nvo arquivo das etiqueta
        caminho = f"Etiqueta_{odp_limpo}_{operario}.xlsx"

        caminho_saida = os.path.join("temporario", caminho)

        planilha.save(caminho_saida)

        return caminho_saida

def salvar_rastreabilidade(rastreabilidade):

        caminho = "temporario/rastreabilidade.json"

        with open( caminho, "w", encoding= "utf-8") as arquivo:
                json.dump(rastreabilidade,arquivo,ensure_ascii=False,indent=4)

def carregar_rastreabilidade():

    caminho = "temporario/rastreabilidade.json"

    print("BUSCANDO RASTREABILIDADE EM:", os.path.abspath(caminho))
    print("ARQUIVO EXISTE?", os.path.exists(caminho))

    if not os.path.exists(caminho):
        return {}

    with open(caminho, "r", encoding="utf-8") as arquivo:
        return json.load(arquivo)

def buscar_rastreabilidade(identificador):

    rastreabilidade = carregar_rastreabilidade()

    if identificador not in rastreabilidade:
        return None
       
    return rastreabilidade[identificador]

def formatar_data_aba(data):

    return(
              data[:2] + "-" +
              data[2:4] + "-" +
              data[4:]
       )

def localizar_aba(dados):

      data = formatar_data_aba(dados["data"])
      print(data)

      turno = dados["turno"]

      nome_arquivo = f"{data}_{turno}.xlsx"

      caminho_saida = os.path.join(f"temporario",nome_arquivo)

      print(f"Buscando arquivo em: {os.path.abspath(caminho_saida)}")

      if not os.path.exists(caminho_saida):
        print("ARQUIVO NÃO ENCONTRADO")
        return None

      planilha = load_workbook(caminho_saida)

      print("ODP QUE ESTOU PROCURANDO:", repr(dados["odp"]))

      aba = planilha.active

      for aba in planilha.worksheets:

        for linha in range(6, aba.max_row + 1):

                valor_odp = aba[f"D{linha}"].value

                print(
                        "LINHA:", linha,
                        "| ODP NA PLANILHA:", repr(valor_odp)
                )

                if str(aba[f"D{linha}"].value) == str(dados["odp"]).strip():
                        print("ODP encontrada na linha:", linha)

                        return {
                                "arquivo": caminho_saida,
                                "aba": aba.title,
                                "linha": linha
                                }
        print("ODP não encontrada")
        return None

def localizar_por_identificador(identificador):

	dados = buscar_rastreabilidade(identificador)

	if dados is None:
		return None

	localizacao = localizar_aba(dados)

	if localizacao is None:
		return None

	return localizacao

	