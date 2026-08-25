from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os
from openpyxl.drawing.image import Image
from datetime import datetime

ALINHAMENTO_PADRAO = Alignment(horizontal="center", vertical="center",wrap_text=True)

FONTE_PADRAO = Font(name="Arial", size=11)

def agrupar_por_operador(ordens):

    grupos = {}

    for ordem in ordens:
        operador = ordem["operador"]

        if operador not in grupos:
            grupos[operador] = []

        grupos[operador].append(ordem)

    return grupos

def selecionar_apontamento(turno, operador,ordens_operador):

    if turno == "Manhã":
        caminho_modelo = "modelos/apontamento_manha.xlsx"
    elif turno == "Noite":
        caminho_modelo = "modelos/apontamento_noite.xlsx"
    else:
        raise Exception(f"{turno} é um turno inválido   ")

    return preencher_apontamento(caminho_modelo, operador, ordens_operador)

def preencher_apontamento(caminho_modelo,operador,ordens_operador):

    planilha = load_workbook(caminho_modelo)

    aba = planilha.active

    primeira_ordem = ordens_operador[0]

    linha = 4

    aba["M2"] = primeira_ordem["data"]
    aba["M2"].number_format = "dd/mm/yyyy"
    aba["M2"].font = FONTE_PADRAO
    aba["M2"].alignment = ALINHAMENTO_PADRAO

    aba["O2"] = primeira_ordem["operador"]
    aba["O2"].font = FONTE_PADRAO
    aba["O2"].alignment = ALINHAMENTO_PADRAO

    aba["Q2"] = primeira_ordem["maquina"]
    aba["Q2"].font = FONTE_PADRAO
    aba["Q2"].alignment = ALINHAMENTO_PADRAO

    for ordem in ordens_operador:
        aba[f"A{linha}"] = ordem["numero_pedido"]
        aba[f"B{linha}"] = ordem["odp"]
        aba[f"C{linha}"] = ordem["cliente"]
        aba[f"E{linha}"] = ordem["padrao"]

        aba[f"A{linha}"].font = FONTE_PADRAO
        aba[f"B{linha}"].font = FONTE_PADRAO
        aba[f"C{linha}"].font = FONTE_PADRAO
        aba[f"E{linha}"].font = FONTE_PADRAO

        aba[f"A{linha}"].alignment = ALINHAMENTO_PADRAO
        aba[f"B{linha}"].alignment = ALINHAMENTO_PADRAO
        aba[f"C{linha}"].alignment = ALINHAMENTO_PADRAO
        aba[f"E{linha}"].alignment = ALINHAMENTO_PADRAO

        linha += 1

    apontamento_arquivo = f"Apontamento_{operador}.xlsx"

    caminho = os.path.join("temporario", apontamento_arquivo)

    planilha.save(caminho)

    planilha.close()

    return caminho

def localizar_odp(aba, odp):

# Aqui ele analisará cada linha desde a 1 até o final, mas devido ao python considerar que ele lerá somente até o valor antes do último colocamos o +1
    for linha in range(1, aba.max_row + 1):

# Se o valor da célula for igual a odp retornará linha
        if aba[f"D{linha}"].value == odp:
            return linha

# Se não retornará nada 
    return None

def preencher_expedição(aba,odp,numero_pallet,peso_total,peso_liquido,op_material,op_tubete):

    linha = localizar_odp(aba,odp)

    if linha is None:
        print(f"ODP não encontrada: {odp}")
        return False
    
    aba[f'E{linha}'] = numero_pallet
    aba[f'F{linha}'] = peso_total
    aba[f'G{linha}'] = peso_liquido
    aba[f'H{linha}'] = op_material
    aba[f'I{linha}'] = op_tubete

    print(f"ODP {odp} encontrada na linha {linha}")

    return True

def preencher_odps(ordens):

    primeira_ordem = ordens[0]

    planilha = load_workbook("modelos/OdP's de cada odp.xlsx")

# Assim fica especificado pro programa sempre usar de parâmetro a aba com o nome "Modelo"
    aba_modelo = planilha["Modelo"]

    data = primeira_ordem["data"].strftime("%d-%m-%Y")
    nome_aba = f"{data}_{primeira_ordem['turno']}"

    if nome_aba in planilha.sheetnames:
        aba = planilha[nome_aba]
    else:
        aba = planilha.copy_worksheet(aba_modelo)
        aba.title = nome_aba

    logo = Image("imagens/luari_logo_empresa.png")
    logo.width = 150
    logo.height = 60
    aba.add_image(logo, "B2")

# congela o painel a partir ca célula A6
    aba.freeze_panes = "A6"

# Cabeçalho

    aba["D3"] = primeira_ordem["data"]
    aba["D3"].number_format = "dd/mm/yyyy"

    aba["F3"] = primeira_ordem["turno"]

    for celula in ("D3","F3"):
        aba[celula].font = FONTE_PADRAO
        aba[celula].alignment = ALINHAMENTO_PADRAO

    linha = 6

    for ordem in ordens:

        aba[f"C{linha}"] = ordem["numero_pedido"]
        aba[f"D{linha}"] = ordem["odp"]

        # Valores que serão preenchidos na pesagem
        aba[f"E{linha}"] = ""
        aba[f"F{linha}"] = ""
        aba[f"G{linha}"] = ""
        aba[f"H{linha}"] = ""
        aba[f"I{linha}"] = ""

        aba[f"J{linha}"] = ordem["cliente"]
        aba[f"K{linha}"] = ordem["padrao"]
        aba[f"L{linha}"] = ordem["filme"]
        aba[f"M{linha}"] = ordem["peso_tubete"]
        aba[f"N{linha}"] = ordem.get("observacao","")
        aba[f"O{linha}"] = ordem["operador"]
        aba[f"P{linha}"] = ordem["maquina"]

        # Dados de rastreabilidade, para serem registrados no histórico
        aba[f"Q{linha}"] = ""
        aba[f"R{linha}"] = ""
        aba[f"S{linha}"] = ""
        aba[f"T{linha}"] = ordem["identificador"]
        aba[f"U{linha}"] = ""
        aba[f"V{linha}"] = ""
        aba[f"W{linha}"] = ""

# Essas letras todas são todas as colunas que tem informação na nossa tabela na qual aplicaremos alguma mudança
        for coluna in "CDEFGHIJKLMNOPQRSTUVW":

            aba[f"{coluna}{linha}"].font = FONTE_PADRAO
            aba[f"{coluna}{linha}"].alignment = ALINHAMENTO_PADRAO

    # OBS em vermelho
        aba[f"N{linha}"].font = Font(name="Arial", size=11, color="FF0000", bold=True)

        linha += 2

    linha_historico = linha + 1

# Remove a aba_modelo antes de salvar
    if aba_modelo.title in planilha.sheetnames:
        planilha.remove(aba_modelo)

    caminho_saida = os.path.join("temporario", f"{data}_{primeira_ordem['turno']}.xlsx")

    planilha.save(caminho_saida)
    planilha.close()

    print("SALVANDO O ARQUIVO:", os.path.abspath(caminho_saida))
    return caminho_saida

def localizar_fim_odps(aba):

# É igual a 0 pois caso a planilha esteja vazia, ou seja, não tenha linhas preenchidas, ele retornará 0
    ultima_linha = 0

# Começa da linha 6 e daí por diante olha todas as linhas preenchidas
    for linha in range(6, aba.max_row + 1):

# Se a ODP ain da estiver preenchida atualiza o valor da ultima_linha,quandfo retornar vazio,o ultimo valor será o que retornará para a próxima função
        if aba[f"D{linha}"].value is not None:
            ultima_linha = linha

    return ultima_linha

def atualizar_odp(caminho_arquivo,nome_aba,linha,identificador,dados):

    planilha = load_workbook(caminho_arquivo)

    aba = planilha[nome_aba]

    campos = {
        "numero_pallet": 'E',
        "peso_liquido": "F",
        "peso_total": "G",
        "op_material": "H",
        "op_tubete": "I"
    }   

    houve_alteracao = False 

# Item faz uma lista onde campo é a chave e coluna valor
    for campo,coluna in campos.items():

# O .get() irá procurar se algum dos valores de campo está dados, se não tiver retornará None
        novo_valor = dados.get(campo)

# Se o operador não preencheu o campo ou deixou em branco, não altera nada, ele continua
        if novo_valor in (None,""):
            continue
# Acessa a célula exata da aba principal e lê o cionteúdo del através do .value()
        valor_anterior = aba[f"{coluna}{linha}"].value 

# Se os valores forem iguais o sistema continua, isso evita que o programa gaste processamento alterando valores iguais e acrescentando linhas de histórico inúteis 
        if str(valor_anterior) == str(novo_valor):
            continue

# Após passar pela filtragem as células da aba principal recebe o novo valor
        aba[f"{coluna}{linha}"] = novo_valor
        houve_alteracao = True

# Aqui estamos usando a função pra descobrir qual a ultima linha das ODP's pra somar mais 2 pra ter o espaço de 1 linha,e começar a gravar o histórico
        linha_historico = localizar_fim_odps(aba) + 1   

        while aba[f"Q{linha_historico}"].value is not None:
            linha_historico += 1
            
        agora = datetime.now()

        aba[f"Q{linha_historico}"] = aba[f"D{linha}"].value
        aba[f"R{linha_historico}"] = agora.strftime("%d/%m/%Y")
        aba[f"S{linha_historico}"] = agora.strftime("%H:%M:%S")
        aba[f"T{linha_historico}"] = identificador
        aba[f"U{linha_historico}"] = dados.get("acao", "pesagem")
        aba[f"V{linha_historico}"] = valor_anterior
        aba[f"W{linha_historico}"] = novo_valor

        for coluna_historico in "QRSTUVW":
            aba[f"{coluna_historico}{linha_historico}"].font = FONTE_PADRAO
            aba[f"{coluna_historico}{linha_historico}"].alignment = ALINHAMENTO_PADRAO

    if houve_alteracao:
        planilha.save(caminho_arquivo)

    planilha.close()

    return houve_alteracao

def registrar_historico(caminho_arquivo,identificador,odp,acao,valor_anterior,novo_valor):

    planilha = load_workbook(caminho_arquivo)

    aba = planilha.active

# Essa linha vizualiza qual a última linha preenchida da tabela +1 pra garantir que comece na linha em branco
    linha = aba.max_row + 1

    agora = datetime.now().strftime("%H:%M:%S")

    aba[f"D{linha}"] = agora
    aba[f"E{linha}"] = identificador
    aba[f"F{linha}"] = odp
    aba[f"G{linha}"] = acao
    aba[f"H{linha}"] = valor_anterior
    aba[f"I{linha}"] = novo_valor

    planilha.save(caminho_arquivo)
    planilha.close()
# Como funciona o Apontamento

# app.py
#   │
#   │ converter_google_sheets(link)
#   ▼
# google_sheet.py
#   │
#   ├── validar_link()
#   │
#   ├── baixar_planilha()
#   │
#   ├── abrir_planilha()
#   │
#   ├── identificar_blocos()
#   │       │
#   │       └── retorna → ordens
#   │
#   ├── selecionar_interpretador()
#   │       │
#   │       └── retorna informações interpretadas
#   │
#   ├── preencher_planilha()
#   │       │
#   │       └── cria arquivo de rastreabilidade
#   │
#   ├── agrupar_por_operador(ordens)
#   │       │
#   │       └── retorna → grupos
#   │
#   └── para cada grupo:
#           │
#           ▼
#     selecionar_apontamento()
#           │
#           ├── verifica turno
#           │
#           ├── escolhe modelo
#           │
#           │   ├── Manhã → apontamento_manha.xlsx
#           │   └── Noite → apontamento_noite.xlsx
#           │
#           ▼
#     preencher_apontamento()
#           │
#           ├── load_workbook()
#           │
#           ├── pega aba
#           │
#           ├── pega primeira ordem
#           │
#           ├── preenche data
#           ├── preenche operador
#           ├── preenche máquina
#           │
#           ├── for ordem in ordens_operador
#           │       │
#           │       ├── número pedido
#           │       ├── ODP
#           │       ├── cliente
#           │       └── padrão
#           │
#           ├── planilha.save()
#           │
#           └── retorna nome do arquivo
#           │
#           ▼
#     arquivos.append(arquivo)
#           │
#           ▼
#     criar_zip(arquivos)
#           │
#           ▼
#     retorna arquivo_zip
#           │
#           ▼
#     app.py
#           │
#           ▼
#     send_file(arquivo_saida)
#           │
#           ▼
#     USUÁRIO RECEBE O ZIP