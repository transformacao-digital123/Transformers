# OpenPyXL é uma biblioteca especializada em ler e escrever
# arquivos do Excel preservando sua estrutura.  
from openpyxl import load_workbook
import os

from servicos.planilha import preencher_planilha
from servicos.utilirarios import criar_zip
from servicos.interpretador import selecionar_interpretador
from servicos.apontamento import agrupar_por_operador, selecionar_apontamento, preencher_odps
from servicos.qrcode import preencher_etiqueta,salvar_rastreabilidade

# Ela Conversa com servidores na internet, desde: acessar sites;baixar imagens; baixar PDFs; baixar Excel; acessar APIs.
import requests

def converter_google_sheets(link):

    rastreabilidade = {}

    os.makedirs("temporario", exist_ok=True)

    for nome in os.listdir("temporario"):
         caminho = os.path.join("temporario", nome)

         if os.path.isfile(caminho):

              try:
                os.remove(caminho)

              except PermissionError:
                   print(f"Não foi possível apagar {nome}")

# Valida de o link é do google docs ou não
    def validar_link(link):
        if "docs.google.com/spreadsheets" not in link:
                raise Exception("O link inserido não é um Google-Sheets")

    def baixar_planilha(link):

# Substitui o final do link e o transforma em um arquivo baixável no navegador
        link_xlsx = link.replace("/edit?", "/export?format=xlsx&")

# requests.get(...): "Vá até esse endereço e me traga a resposta, um objeto que contém tudo que o servidor respondeu. 
# Enquanto isso o "resposta.content" irá guardar em bytes a informação no computador até ela ser usada"
        resposta = requests.get(link_xlsx)

# Variável para que possamos sempre usar o nome que será retornada, quando quisermos só alterando 1 linha
        caminho_arquivo = "planilha.xlsx"

# Crie (ou abra) um arquivo chamado planilha.xlsx, no modo de escrita binária, e quando eu terminar, feche-o automaticamente.
        with open(caminho_arquivo, "wb") as arquivo:
             arquivo.write(resposta.content)

        return caminho_arquivo

    def abrir_planilha(caminho_arquivo):
         
# Abre a planilha salva no computador e carrega seu conteúdo para a memória, ou seja, abre a planilha salva no computador         
        planilha = load_workbook(caminho_arquivo)

        return planilha
         
    def identificar_blocos(aba):

        maquina = ""
        operador = ""
        turno = ""

        estado ="procurando_bloco"

# Para armazenar as informações do dicionário "dados"
        ordens=[]

        colunas = {}

        MAPEAMENTO = {
            "data": "DATA",
            "numero_pedido": "PEDIDO",
            "odp": "OdP",
            "cliente": "CLIENTE",
            "padrao": ["PADRÃO", "LARGURA  X MICRA"]
        }

# O .iter_rows() lê linha por linha do conteúdo que está conectado
        for linha in aba.iter_rows(): 

            try:

                valores = []

            except Exception as erro:

                raise

# Percorre as células da linha  
            for celula in linha:
                if celula.value is not None:

# O .append acrescenta na lista Valores, e o .strip separa o conteúdo por espaços 
                        valores.append(str(celula.value).strip())

# o len lê a quantidade, se a quantidade de celulas preenchidas for 1 continuará para descobrir a máquina e o operador
            if len(valores) == 1:
                    
        # Guarda o primeiro valor da célula na variável texto
                    texto = valores[0]

        # Primeiro verifica se é o título da ODP
                    if texto.startswith("Ordem De Produção"):

                        if "Noite" in texto:
                            turno = "Noite"
                        elif "Manhã" in texto:
                            turno = "Manhã"

        # Só depois verifica máquina-operador
                    elif " - " in texto:

                        maquina, operador = texto.split(" - ", 1)

                        maquina = maquina.strip()
                        operador = operador.strip()

# Descobrir o cabeçalho
            if "DATA" in valores and "PEDIDO" in valores and "OdP" in valores:
                estado = "lendo_ordens"


                if "PADRÃO" in valores:
                     origem = "PADRÃO"
                     coluna_padrao = "PADRÃO"
                     
                elif "LARGURA  X MICRA" in valores:
                     origem = "LARGURA  X MICRA"
                     coluna_padrao = "LARGURA  X MICRA"

# O indice é a posição, em que coluna está, e a celula a coordenada excel, como A15
                for indice, celula in enumerate(linha):
                    if celula.value:

# Estrutura feita para que caso mudem alguma coluna de lugar essa linha se atualizará sozinha
                        colunas[str(celula.value).strip()] = indice

                continue

            if estado == "lendo_ordens":

                data = linha[colunas["DATA"]].value
                numero_pedido = linha[colunas["PEDIDO"]].value      

                if (data is not None and numero_pedido is not None):

                    dados = {}

# O .items serve para organizar o dicionário em chave e valor
                    for chave_programa,chave_planilha in MAPEAMENTO.items():

                        if chave_programa == "padrao":
                            dados[chave_programa] = linha[colunas[coluna_padrao]].value
                        else:
                            dados[chave_programa] = linha[colunas[chave_planilha]].value

# Aqui não precisa de , pq não é uma string, é uma tupla
                    dados["operador"] = operador
                    dados["maquina"] = maquina
                    dados["observacao"] = ""
                    dados["turno"] = turno

                    dados["origem"] = origem

                    ordens.append(dados)

                if   len(valores) == 1:
                    if valores[0].startswith("OBS: "):
                        ordens[-1]["observacao"] = valores[0].replace("OBS: ","")

        return ordens

# Esse está aqui apenas para orietação
    validar_link(link)

    caminho_arquivo = baixar_planilha(link)
    planilha = abrir_planilha(caminho_arquivo)

# Seleciona a aba ativa da planilha
    aba = planilha.active
    ordens = identificar_blocos(aba)

    planilha.close()

    os.remove(caminho_arquivo)

    arquivos = []

    print(len(ordens))
    for indice,ordem in enumerate(ordens):
# Função chamada para analisar o padrão, e descobrir o filme e o peso do tubete. Além disso, dentro dela nota-se 2 padrao, o 1° é para encontrar a variável dentro do dicionário e o 2° é para encontrar a coluna caso ela se chame PADRÃO
        try:
                informacoes = selecionar_interpretador(ordem["padrao"], ordem["origem"])
             
                ordem["filme"] = informacoes["filme"]
                ordem["peso_tubete"] = informacoes["peso_tubete"]
                ordem["padrao"] = informacoes["padrao"]

                arquivo = preencher_planilha(ordem, indice)
             
                arquivos.append(arquivo)
        
        except Exception as erro:

            raise

    grupos = agrupar_por_operador(ordens)
                    
    for operador,ordens_operador in grupos.items():

        turno = ordens_operador[0]["turno"]

        arquivo = selecionar_apontamento(turno, operador, ordens_operador)
        arquivos.append(arquivo)

        arquivo_odps = preencher_odps(ordens_operador)
        arquivos.append(arquivo_odps)

        for ordem in ordens_operador:
            arquivo_etiqueta = preencher_etiqueta(ordem,rastreabilidade)
            arquivos.append(arquivo_etiqueta)

    arquivo_zip = criar_zip(arquivos)

    salvar_rastreabilidade(rastreabilidade)

    return arquivo_zip

